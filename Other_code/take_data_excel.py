import pandas as pd
import sys
import os
from openpyxl import load_workbook

# TAKE DATA AMOUNT USING FROM MCG
group_dict = {"PE1": "XUẤT IE1", "PE2": "XUẤT IE2", "PE3": "XUẤT IE3", "PE4": "XUẤT IE4", "PE5": "XUẤT PEM"} 
output_folder = r"X:\CY26\Safety Stock\Data\NEW"
PE1_df = pd.DataFrame(columns=['part_code'])
PE2_df = pd.DataFrame(columns=['part_code'])
PE3_df = pd.DataFrame(columns=['part_code'])
PE4_df = pd.DataFrame(columns=['part_code'])
PE5_df = pd.DataFrame(columns=['part_code'])
result_dict = {"PE1": PE1_df, "PE2": PE2_df, "PE3": PE3_df, "PE4": PE4_df, "PE5": PE5_df}

def read_excel_file():
    try:
        for year in range(2024, 2027):
            year_2 = str(year)[-2:]
            folder_path  = rf"\\172.30.73.156\mcg\QUAN LY CCDC&PTTT\Bao Cao Kiem Soat CCDC&PTTT\Bao cao da gui\Bao cao {year}\PE"
            for month in range(1, 13):
                if year == 2026 and month > 6:
                   break  # Stop processing months after June 2026
                if year == 2024:
                    file_name = f"Bao Cao {month}'{year_2}.xlsx"
                elif year == 2025:
                    file_name = f"Bao Cao PE {month}'{year_2}.xlsx"
                elif year == 2026:
                    file_name = f"Bao Cao(HT) {month:02d}'{year_2}.xlsx"
                file_path = os.path.join(folder_path, file_name)
                for key, value in group_dict.items():
                    try:
                        df = pd.read_excel(file_path, sheet_name=value,skiprows=15,usecols="B,O")
                    except Exception as e:
                        print(f"Error reading sheet {value} in file {file_name}: {e}")
                        continue
                    df.columns = ['part_code',f"T{month}-{year_2}"] 
                    df['part_code'] = df['part_code'].astype(str).str.strip()
                    df[f"T{month}-{year_2}"] = pd.to_numeric(df[f"T{month}-{year_2}"], errors='coerce').fillna(0).astype(int)
                    if result_dict[key].empty:
                        result_dict[key] = df
                    else:
                        result_dict[key] = result_dict[key].merge(df, on='part_code', how='outer')
        for key in result_dict:
            result_dict[key].fillna(0, inplace=True)
            result_dict[key].to_excel(os.path.join(output_folder, f"{key}_result.xlsx"), index=False)
    except Exception as e:
        print(f"Error reading the Excel file: {e}")
        return None

#=========================================================================================================================
#=========================================================================================================================
#=========================================================================================================================

# TAKE DATA OUTPUT FROM PRODUCTION

# month_dict = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun", 7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}
# def read_excel_file():
#     result_df = pd.DataFrame(columns=['SBU'])
#     try:
#         for year in range(2025, 2026):
#             year_2 = str(year)[-2:]
#             folder_path  = rf"\\172.30.73.156\factorydata\FY2021-PO1\PO2\{year}"
#             for month in range(1, 13):
#                 file_name = f"PO2-Target(Pcs.hr) {month_dict[month]}.{year_2}.xlsx"
#                 print(f"Processing file: {file_name}")
#                 file_path = os.path.join(folder_path, file_name)
#                 try:
#                     wb = load_workbook(file_path, read_only=True, data_only=True)
#                     ws = wb["Summary"]

#                     rows = []
#                     for i, row in enumerate(ws.iter_rows(min_row=4, max_col=5, values_only=True)):  # skiprows=3 → min_row=4 hoặc 5
#                         sbu = row[1]  # cột B (index 1)
#                         qty = row[4]  # cột E (index 4)
#                         if str(sbu).strip() == "Total" or str(sbu).strip() == "NEW":
#                             break
#                         rows.append((sbu, qty))

#                     wb.close()
#                     df = pd.DataFrame(rows, columns=["SBU", f"T{month}-{year_2}"])
#                     mask = df["SBU"] == "Total"
#                     df = df.iloc[:mask.idxmax()] if mask.any() else df
#                     df["SBU"] = df["SBU"].astype(str).str.strip()
#                     df[f"T{month}-{year_2}"] = pd.to_numeric(df[f"T{month}-{year_2}"], errors='coerce').fillna(0).astype(int)
#                     if result_df.empty:
#                         result_df = df
#                     else:
#                         result_df = result_df.merge(df, on='SBU', how='outer')
#                 except Exception as e:
#                     print(f"Error reading sheet Summary in file {file_name}: {e}")
#                     continue
#         result_df.fillna(0, inplace=True)
#         output_folder = r"X:\CY26\Safety Stock\Data"
#         result_df.to_excel(os.path.join(output_folder, f"PO2_Target_{year}.xlsx"), index=False)
#     except Exception as e:
#         print(f"Error reading the Excel file: {e}")
#         return None

# def read_excel_file():
#     folder_path = r"X:\Customer Complain\Auto\LOG"
#     list_file = os.listdir(folder_path)
#     df_final = pd.DataFrame()
#     for file in list_file:
#         file_path = os.path.join(folder_path, file)
#         df = pd.DataFrame()
#         if file.endswith("csv"):
#             try:
#                 # Kiểm tra nếu file thực ra là HTML giả dạng .xls
#                 with open(file_path, "rb") as f:
#                     header = f.read(20)
#                 if b"<html" in header.lower() or header.startswith(b"\xef\xbb\xbf"):
#                     tables = pd.read_html(file_path, encoding="utf-8")
#                     df = tables[0]
#                     # Flatten MultiIndex columns nếu có
#                     if isinstance(df.columns, pd.MultiIndex):
#                         df.columns = [" ".join(str(c) for c in col).strip() for col in df.columns]
#                     df["Lot No."] = file.split(".")[0]
#                 else:
#                     for enc in ("utf-8", "cp932", "shift_jis", "latin-1"):
#                         try:
#                             df = pd.read_csv(file_path, encoding=enc, skiprows=1, sep=None, engine="python")
#                             break
#                         except (UnicodeDecodeError, ValueError):
#                             continue
#                     df["Lot No."] = file.split(".")[0]
#             except Exception as e:
#                 print(f"Error reading {file}: {e}")
#         df_final = pd.concat([df_final, df], ignore_index=True)
#     output_folder = r"F:\2173452100291"
#     df_final.to_excel(os.path.join(output_folder, "Final_Result_MA16_Feb.xlsx"), index=False)

if __name__ == "__main__":
    result = read_excel_file()